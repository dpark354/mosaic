#%%
#! python3.9 -m pip install --upgrade pip
#! pip install modin[ray]
#! pip install openpyxl
#! pip install pandas
#! pip install tqdm
#! pip install numpy
#! pip install seaborn
#! pip install sklearn
#%%

! pip install 'modin[ray]'
#%%
import openpyxl as pyxl
from openpyxl import load_workbook
#import modin.pandas as pd
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from tqdm.notebook import tqdm, tnrange
import ray
#ray.init(runtime_env={'env_vars': {'__MODIN_AUTOIMPORT_PANDAS__': '1'}})
#tqdm.pandas()


def read_file(file_, tab_):
    ##read file and
    wb=pyxl.load_workbook(file_, read_only=True)
    ws=wb[tab_]
    d=pd.DataFrame(ws.values)
    #d=pd.read_excel(file_,tab_)

    d.columns=d.iloc[0,:]
    d=d.iloc[1:,:]
    d=d.reset_index(drop=True)
    return d

def read_file_csv(file_, tab_):
    ##read file and
    d=pd.read_csv(file_)
    #d=pd.read_excel(file_,tab_)

    #d.columns=d.iloc[0,:]
    #d=d.iloc[1:,:]
    #d=d.reset_index(drop=True)
    return d

def date_sortable(create_date, delivered_date, modified_date):
    ## take the first available date, modified trumps delivered, trumps created

    if len(modified_date)>3:
        return modified_date
    elif len(delivered_date)>3:
        return delivered_date
    else:
        return create_date

def month_year(date_text):
    ##only works with a text field, but it pulls the 7 digits

    return date_text[:7]

def map_bucket(yyyy_mm):
    yyyy=int(yyyy_mm[:4])
    mm=int(yyyy_mm[:-2])
    if mm<4:
        Q=1
    elif mm<7:
        Q=2
    elif mm<10:
        Q=3
    else:
        Q=4
    return (f'{yyyy}Q{Q}')

def softmax(val_,threshold_):
    try:
        if float(val_)>threshold_:
            return 1
        else:
            return 0
    except:
        return 0

def clean_name(d):
    try:
        clean1=d.split('@')
        clean2=clean1[0].split('[')
        clean3=clean2[0].split('(')
        clean4=clean3[0].replace('\.',' ')
        clean5=clean4.strip()
        d=clean5
        at_=clean1[1]
    except:
        d=""
        at_=""
    return d

def at_name(d):
    try:
        clean1=d.split('@')
        clean2=clean1[0].split('[')
        clean3=clean2[0].split('(')
        clean4=clean3[0].replace('\.',' ')
        clean5=clean4.strip()
        d=clean5
        at_=clean1[1]
    except:
        d=""
        at_=""
    return at_

def make_float(x):
    try:
        y=float(x)
    except:
        y=-1
    return y
#%%
##input file names, change them to the ones being processed

wb_original='30155_PII_20230215.csv'
wb_prefix=wb_original[:17]
initial_tab='30155_2023-02-07_results'

d_scratch1=read_file_csv(wb_original, initial_tab)
#%%
col=d_scratch1.columns.values
phi=""
pii=""
for i in col:
    l=i.split('_')
    if 'phi' in l:
        phi=i
    if 'pii' in l:
        pii=i

##instantiate the columns so they can be updated
d=d_scratch1
d_scratch1['date_sortable']=""
d_scratch1['Year_Month']=""
d_scratch1['phi_7']=d[phi].apply(lambda x: 1 if (0.7<make_float(x)) else 0)
d_scratch1['pii_7']=d[pii].apply(lambda x: 1 if (0.7<make_float(x)) else 0)
d_scratch1['phi_5']=d[phi].apply(lambda x: 1 if (0.5<make_float(x)) else 0)
d_scratch1['pii_5']=d[pii].apply(lambda x: 1 if (0.5<make_float(x)) else 0)
d_scratch1['phi_3']=d[phi].apply(lambda x: 1 if (0.3<make_float(x)) else 0)
d_scratch1['pii_3']=d[pii].apply(lambda x: 1 if (0.3<make_float(x)) else 0)
#val_=0.7
#%%

d=d_scratch1

for d_ in tnrange(len(d_scratch1)):
    d_sort=date_sortable(d.createddate[d_], d.delivereddate[d_], d.modifieddate[d_])
    d.at[d_,'date_sortable']=d_sort

d_scratch1['Year_Month']=d['date_sortable'].apply(lambda x: month_year(x))


d_update=d['owner'].apply(lambda x: clean_name(x))
d_at=d['owner'].apply(lambda x: at_name(x))

d['clean_name']=d_update
d['at']=d_at

d.to_csv(wb_prefix+'_2.csv')
#%%
#%%


#len(d['split_path'])
def split_path(d):
    one_drive_custodian=""
    try:
        u=d.split('/')
    except:
        u=[]
    try:
        u2=u[0]
        if u2 =='~':
            one_drive_custodian=u[1]
            u2=u[2]
    except:
        u2=""
    return u, one_drive_custodian, u2

d['path_split']=d['path'].apply(lambda x: split_path(x) )
#%%
##checkpoint#3
#d['clean_name']=d['owner'].apply(lambda x: clean_name(x))
#dname=d[['owner','clean_name','fullaccount']]
#clean_name=d['clean_name']


#%%


##thesee are the stubs in the name that are almost never valid
def dispo(x):
    List_of_terms=['Admin',\
               'Education',\
                'Zoom', \
                'Summit',\
                'News',\
                'Deal',\
                'Service',\
                'hospi',\
                'ameri',\
                'clinic',\
                'care',\
                'service',\
                '_',\
                'AABB',\
                'AAFP',\
                'AAHAM',\
                'AANN',\
                'AbbVie',\
                'AARC',\
                'Lab',\
                'Path',\
                'Onco',\
                'Nurs',\
                'Socie',\
                'JAMA',\
                'AmTrav',\
                'Amazon',\
                'Train',\
                'conf',\
                'bio',\
                'food',\
                'room',\
                'Team',\
                'Center',\
                'master',\
                'disco',\
                'cardio',\
                'inspir',\
                'credit',\
                'depar',\
                'finan',\
                'share',\
                'support',\
                'help',\
                'care',\
                'reply',\
                'email',\
                'fax',\
                'comp',\
                'The ',\
                'webi',\
                'market',\
                'apic',\
                'lead',\
                'board',\
                'advisor',\
                'date',\
                'pay',\
                'library',\
                'inc',\
                'mem',\
                'college',\
                'uni',\
                'comm',\
                'codi',\
                'code',\
                'medi',\
                'pharm',\
                'from',\
                'golf',\
                'custo',\
                ' at ',\
                'maga',\
                '\&',\
                '\!',\
                'tech',\
                ' in ',\
                'data',\
                'scie',\
                'thermo',\
                'docu',\
                'event',\
                'hello',\
                'employ',\
                'street',\
                'wound',\
                'moni',\
                'covid',\
                'account',\
                'wire',\
                'exc',\
                'facebook',\
                'pain',\
                'ology',\
                'organi',\
                'group',\
                'Today',\
                'Legacy',\
                'PDF',\
                'U S',\
                'and ',\
                'physician',\
                'health',\
                'journal',\
                'academy ',\
                'network',\
                'Microsoft',\
                'apple']





    try:
        if x in List_of_terms:
            return 1
        elif x[-1]=="2":
            return 1
        else:
            return 0
    except:
        return -1
        #if the term is in th elist of names mark it a 1 
d['Dispo']=d['clean_name'].apply(lambda x: dispo(x) )
#dname.to_excel('name_cleanup.xlsx')

#%%

dedupe=pd.DataFrame(d[['clean_name', 'owner','fullaccount','Dispo','at']].drop_duplicates())
#dedupe['Dispo']=dedupe['clean_name'].apply(lambda x: 1 if any(i in x for i in List_of_terms) else 0)
## stop here for tonight

#%%

#look_up=pyxl.load_workbook('name_cleanup2.xlsx')
    #%%
domain_=dedupe[dedupe['at']!=""]['at'].drop_duplicates()
#%%
domain_.reset_index(drop=True)
#%%

domain_.to_excel('domain_cleanup.xlsx')
#%%


d_look_up=pd.read_csv('domain_cleanup.csv')

#%%
d_look_up=d_look_up[['at','disposition']]
#%%


d_joined=d.join(d_look_up.set_index('at'), on='at')

#%%
d_joined[d_joined['disposition'].isna()]
# test to confirm names are working in the beginning and the end

#d_joined.to_excel('join_mosaic.xlsx')
#%%